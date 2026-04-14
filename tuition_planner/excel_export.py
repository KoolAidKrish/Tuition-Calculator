from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="203A43")
ACCENT_FILL = PatternFill("solid", fgColor="2C5364")
WHITE_FONT = Font(color="FFFFFF", bold=True)


def _style_sheet_header(worksheet, title: str) -> None:
    worksheet["A1"] = title
    worksheet["A1"].fill = ACCENT_FILL
    worksheet["A1"].font = WHITE_FONT


def _auto_size_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(length + 2, 28)


def export_workbook(
    output_path: str | Path,
    summary: dict[str, object],
    tuition_table,
    savings_table,
    spending_summary=None,
) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _style_sheet_header(summary_sheet, "Scenario Summary")

    for index, (key, value) in enumerate(summary.items(), start=3):
        summary_sheet[f"A{index}"] = key.replace("_", " ").title()
        summary_sheet[f"B{index}"] = value
        summary_sheet[f"A{index}"].fill = HEADER_FILL
        summary_sheet[f"A{index}"].font = WHITE_FONT

    tuition_sheet = workbook.create_sheet("Tuition Forecast")
    _style_sheet_header(tuition_sheet, "Forecasted Tuition")
    tuition_headers = list(tuition_table.columns)
    for index, header in enumerate(tuition_headers, start=1):
        cell = tuition_sheet.cell(row=3, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    for row_index, row in enumerate(tuition_table.itertuples(index=False), start=4):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, (int, float)):
                value = float(value)
            tuition_sheet.cell(row=row_index, column=column_index, value=value)

    savings_sheet = workbook.create_sheet("Savings Plan")
    _style_sheet_header(savings_sheet, "Savings Projection")
    savings_headers = list(savings_table.columns)
    for index, header in enumerate(savings_headers, start=1):
        cell = savings_sheet.cell(row=3, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
    for row_index, row in enumerate(savings_table.itertuples(index=False), start=4):
        for column_index, value in enumerate(row, start=1):
            if hasattr(value, "to_pydatetime"):
                value = value.to_pydatetime()
            elif isinstance(value, (int, float)):
                value = float(value)
            savings_sheet.cell(row=row_index, column=column_index, value=value)

    if spending_summary is not None:
        spending_sheet = workbook.create_sheet("Spending Insights")
        _style_sheet_header(spending_sheet, "Spending Metrics")
        for index, (key, value) in enumerate(spending_summary.items(), start=3):
            spending_sheet[f"A{index}"] = key.replace("_", " ").title()
            spending_sheet[f"B{index}"] = value
            spending_sheet[f"A{index}"].fill = HEADER_FILL
            spending_sheet[f"A{index}"].font = WHITE_FONT
        _auto_size_columns(spending_sheet)

    for sheet in workbook.worksheets:
        _auto_size_columns(sheet)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
